import openpyxl

def dataGenerator():
    #li = [['uname1','pass1'],['uname2','pass2'],['uname3','pass3']] Reading data directly and not from excel
    #return li
    wk = openpyxl.load_workbook("C:\\Users\\Jashan\\Documents\\Python\\TData.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    li=[]
    li1=[]#creating a list inside a list as done above in direct method
    for i in range(1,r+1):
        li1=[]
        un = sh.cell(i,1)
        up = sh.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,up.value)
        li.insert(i-1,li1)
    print(li)
    return li